#!/usr/bin/env python3
"""Audit the committee data: sheet completeness + Burget(ová) disambiguation."""
import json, re, sys, unicodedata
from pathlib import Path
from collections import Counter, defaultdict
import warnings; warnings.filterwarnings("ignore")
import openpyxl

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "materials" / "MSZ 2026 ALL KOMISE.xlsx"
sys.path.insert(0, str(Path(__file__).resolve().parent))
from build import member_key, fold

wb = openpyxl.load_workbook(XLSX, data_only=True)
print("SHEETS:", wb.sheetnames)

def rows_of(sheet):
    ws = wb[sheet]
    out = []
    session = "?"
    for r in ws.iter_rows(values_only=True):
        r = list(r) + [None] * 5
        if r[0] and str(r[0]).strip():
            session = str(r[0]).strip()
        member = str(r[1]).strip() if r[1] else ""
        c2 = str(r[2]).strip() if r[2] is not None else ""
        course = str(r[3]).strip() if r[3] else ""
        text = str(r[4]).strip() if r[4] else ""
        if member or course or text:
            out.append({"session": session, "member": member, "c2": c2, "course": course, "text": text})
    return out

def fp(rec):
    # fingerprint by course + first 60 chars of (text or c2), accent/space-insensitive
    body = (rec["text"] or rec["c2"])[:60]
    return (fold(rec["course"]), re.sub(r"\s+", " ", fold(body)).strip())

l1 = rows_of("List 1")
s24 = rows_of("2024") if "2024" in wb.sheetnames else []
s22 = rows_of("2022") if "2022" in wb.sheetnames else []
print(f"\nList 1 filled rows : {len(l1)}")
print(f"'2024' filled rows : {len(s24)}")
print(f"'2022' filled rows : {len(s22)}")

# overlap: which '2024' records are NOT present in List 1?
l1_fps = Counter(fp(r) for r in l1)
uniq24 = [r for r in s24 if l1_fps[fp(r)] == 0]
print(f"\n'2024' records NOT found in List 1 (by course+text fp): {len(uniq24)}")
for r in uniq24[:25]:
    print(f"   [{r['member'][:18]:18s}] {r['course']:5s} {(r['text'] or r['c2'])[:55]}")
if len(uniq24) > 25:
    print(f"   … +{len(uniq24)-25} more")

# also: '2024' records whose member differs even if text matches?
print(f"\n'2022' sample:")
for r in s22[:5]:
    print("  ", r)

# ── Burget / Burgetová in the raw data ──
print("\n── Burget* raw strings across List 1 (+2024) ──")
allrows = l1 + s24
bagg = defaultdict(lambda: {"count": 0, "courses": Counter(), "firsts": Counter(), "keys": Counter()})
for r in allrows:
    if "burget" in fold(r["member"]):
        key, first = member_key(r["member"])
        b = bagg[r["member"]]
        b["count"] += 1
        if r["course"]: b["courses"][r["course"]] += 1
        b["keys"][key] += 1
        if first: b["firsts"][first] += 1
for raw, b in sorted(bagg.items(), key=lambda x: -x[1]["count"]):
    key = b["keys"].most_common(1)[0][0]
    print(f"  {b['count']:3d}  raw={raw[:34]:34s} -> key={key}  courses={dict(b['courses'].most_common(6))}")

# bare "Burget" by course (to design L/R split)
print("\n── bare 'Burget' (no first name) records by course ──")
bare = Counter()
for r in allrows:
    if fold(r["member"]) == "burget":
        bare[r["course"]] += 1
print("  ", dict(bare.most_common(20)))

# ── current output JSON ──
out = ROOT / "public" / "repos" / "fit-msz.json"
if out.exists():
    d = json.loads(out.read_text(encoding="utf-8"))
    print(f"\n── current fit-msz.json: {len(d['members'])} members, {len(d['records'])} records ──")
    for m in d["members"]:
        if "burget" in m["key"]:
            print(f"   key={m['key']:12s} display={m['display']:24s} count={m['count']} aliases={m['aliases']}")
    # any records whose course/member mention burgetova but mapped elsewhere?
    bv = [r for r in d["records"] if r["memberKey"] == "burgetova"]
    print(f"   records with memberKey=='burgetova': {len(bv)}")
wb.close()
